# -*- coding: utf-8 -*-
########################################################
# CMeIE 评价脚本文件
########################################################
"""
This module to calculate precision, recall and f1-value 
of the predicated results.
"""
import sys
import json
import os
import zipfile
import traceback
import argparse
from openpyxl import Workbook

SUCCESS = 0
FILE_ERROR = 1
NOT_ZIP_FILE = 2
ENCODING_ERROR = 3
JSON_ERROR = 4
SCHEMA_ERROR = 5
ALIAS_FORMAT_ERROR = 6

CODE_INFO = {
    SUCCESS: 'success',
    FILE_ERROR: 'file is not exists',
    NOT_ZIP_FILE: 'predict file is not a zipfile',
    ENCODING_ERROR: 'file encoding error',
    JSON_ERROR: 'json parse is error',
    SCHEMA_ERROR: 'schema is error',
    ALIAS_FORMAT_ERROR: 'alias dict format is error'
}


def del_bookname(entity_name):
    """delete the book name"""
    if entity_name.startswith(u'《') and entity_name.endswith(u'》'):
        entity_name = entity_name[1:-1]
    return entity_name


def check_format(line):
    """检查输入格式是否错误"""
    ret_code = SUCCESS
    json_info = {}
    try:
        line = line.strip()
    except:
        ret_code = ENCODING_ERROR
        return ret_code, json_info
    try:
        json_info = json.loads(line)
    except:
        ret_code = JSON_ERROR
        return ret_code, json_info
    if 'text' not in json_info or 'spo_list' not in json_info:
        ret_code = SCHEMA_ERROR
        return ret_code, json_info
    required_key_list = ['subject', 'predicate', 'object']
    for spo_item in json_info['spo_list']:
        if type(spo_item) is not dict:
            ret_code = SCHEMA_ERROR
            return ret_code, json_info
        if not all(
                [required_key in spo_item for required_key in required_key_list]):
            ret_code = SCHEMA_ERROR
            return ret_code, json_info
        if not isinstance(spo_item['subject'], str) or \
                not isinstance(spo_item['object'], dict):
            ret_code = SCHEMA_ERROR
            return ret_code, json_info
    return ret_code, json_info


# def _parse_structured_ovalue(json_info):
#     spo_result = []
#     for item in json_info["spo_list"]:
#         s = del_bookname(item['subject'].lower())
#         o = {}
#         for o_key, o_value in item['object'].items():
#             o_value = del_bookname(o_value).lower()
#             o[o_key] = o_value
#         spo_result.append({"predicate": item['predicate'], \
#                            "subject": s, \
#                            "object": o})
#     return spo_result
def _parse_structured_ovalue(json_info):
    spo_result = []
    for item in json_info["spo_list"]:
        s = del_bookname(item['subject'].lower())
        o = {}
        for o_key, o_value in item['object'].items():
            o_value = del_bookname(o_value).lower()
            o[o_key] = o_value
        spo_result.append({"predicate": item['predicate'], \
                           "subject": s, \
                           "object": o})
    return spo_result

def load_predict_result(predict_filename):
    """Loads the file to be predicted
        可加载zip文件，已弃用
    """
    predict_result = {}
    ret_code = SUCCESS
    if not os.path.exists(predict_filename):
        ret_code = FILE_ERROR
        return ret_code, predict_result
    try:
        predict_file_zip = zipfile.ZipFile(predict_filename)
    except:
        ret_code = NOT_ZIP_FILE
        return ret_code, predict_result
    for predict_file in predict_file_zip.namelist():
        for line in predict_file_zip.open(predict_file):
            ret_code, json_info = check_format(line)
            if ret_code != SUCCESS:
                return ret_code, predict_result
            sent = json_info['text']
            spo_result = _parse_structured_ovalue(json_info)
            predict_result[sent] = spo_result
    return ret_code, predict_result


def load_test_dataset(golden_filename):
    """load golden file"""
    golden_dict = {}
    ret_code = SUCCESS
    if not os.path.exists(golden_filename):
        ret_code = FILE_ERROR
        return ret_code, golden_dict
    with open(golden_filename, 'r', encoding="utf-8") as gf:
        for line in gf:
            ret_code, json_info = check_format(line)
            if ret_code != SUCCESS:
                return ret_code, golden_dict

            sent = json_info['text']
            spo_result = _parse_structured_ovalue(json_info)
            golden_dict[sent] = spo_result
    return ret_code, golden_dict


def load_alias_dict(alias_filename):
    """load alias dict"""
    alias_dict = {}
    ret_code = SUCCESS
    if alias_filename == "":
        return ret_code, alias_dict
    if not os.path.exists(alias_filename):
        ret_code = FILE_ERROR
        return ret_code, alias_dict
    with open(alias_filename, "r", encoding="utf-8") as af:
        for line in af:
            line = line.strip()
            try:
                words = line.split('\t')
                alias_dict[words[0].lower()] = set()
                for alias_word in words[1:]:
                    alias_dict[words[0].lower()].add(alias_word.lower())
            except:
                ret_code = ALIAS_FORMAT_ERROR
                return ret_code, alias_dict
    return ret_code, alias_dict


def del_duplicate(spo_list, alias_dict):
    """delete synonyms triples in predict result"""
    normalized_spo_list = []
    for spo in spo_list:
        if not is_spo_in_list(spo, normalized_spo_list, alias_dict):
            normalized_spo_list.append(spo)
    return normalized_spo_list


def is_spo_in_list(target_spo, golden_spo_list, alias_dict):
    """target spo是否在golden_spo_list中"""
    if target_spo in golden_spo_list:
        return True
    target_s = target_spo["subject"]
    target_p = target_spo["predicate"]
    target_o = target_spo["object"]
    target_s_alias_set = alias_dict.get(target_s, set())
    target_s_alias_set.add(target_s)
    for spo in golden_spo_list:
        s = spo["subject"]
        p = spo["predicate"]
        o = spo["object"]
        if p != target_p:
            continue
        if s in target_s_alias_set and _is_equal_o(o, target_o, alias_dict):
            return True
    return False


def _is_equal_o(o_a, o_b, alias_dict):
    for key_a, value_a in o_a.items():
        if key_a not in o_b:
            return False
        value_a_alias_set = alias_dict.get(value_a, set())
        value_a_alias_set.add(value_a)
        if o_b[key_a] not in value_a_alias_set:
            return False
    for key_b, value_b in o_b.items():
        if key_b not in o_a:
            return False
        value_b_alias_set = alias_dict.get(value_b, set())
        value_b_alias_set.add(value_b)
        if o_a[key_b] not in value_b_alias_set:
            return False
    return True


def list2dic(spo_list):
    temp_dic = {}
    p_list = []
    for spo in spo_list:
        if spo["predicate"] not in temp_dic:
            temp_dic[spo["predicate"]] = []
        temp_dic[spo["predicate"]].append(spo)
    p_list = list(temp_dic.keys())
    return temp_dic, p_list


def calc_pr(predict_filename, alias_filename, golden_filename):

    """calculate precision, recall, f1
        计算整体的 precision, recall, f1
    """
    ret_info = {}

    #load alias dict {}
    ret_code, alias_dict = load_alias_dict(alias_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info
    #load test golden dataset
    ret_code, golden_dict = load_test_dataset(golden_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info
    #load predict result
    ret_code, predict_result = load_test_dataset(predict_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info

    #evaluation
    correct_sum, predict_sum, recall_sum, recall_correct_sum = 0.0, 0.0, 0.0, 0.0
    for sent in golden_dict:
        # 去重
        golden_spo_list = del_duplicate(golden_dict[sent], alias_dict)
        #
        predict_spo_list = predict_result.get(sent, list())
        normalized_predict_spo = del_duplicate(predict_spo_list, alias_dict)
        #
        recall_sum += len(golden_spo_list)
        predict_sum += len(normalized_predict_spo)
        for spo in normalized_predict_spo:
            if is_spo_in_list(spo, golden_spo_list, alias_dict):
                correct_sum += 1
        for golden_spo in golden_spo_list:
            if is_spo_in_list(golden_spo, predict_spo_list, alias_dict):
                recall_correct_sum += 1
    sys.stderr.write('correct spo num = {}\n'.format(correct_sum))
    sys.stderr.write('submitted spo num = {}\n'.format(predict_sum))
    sys.stderr.write('golden set spo num = {}\n'.format(recall_sum))
    sys.stderr.write('submitted recall spo num = {}\n'.format(
        recall_correct_sum))
    precision = correct_sum / predict_sum if predict_sum > 0 else 0.0
    recall = recall_correct_sum / recall_sum if recall_sum > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) \
            if precision + recall > 0 else 0.0
    precision = round(precision, 4)
    recall = round(recall, 4)
    f1 = round(f1, 4)
    t_lst = ['all', correct_sum, predict_sum, recall_sum, recall_correct_sum, precision, recall, f1]

    return t_lst

def calc_pr_mul(predict_filename, alias_filename, golden_filename):
    """
        计算每一个子关系的 precision, recall, f1
    """
    ret_info = {}

    # load alias dict {}
    ret_code, alias_dict = load_alias_dict(alias_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info
    # load test golden dataset
    ret_code, golden_dict = load_test_dataset(golden_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info
    # load predict result
    ret_code, predict_result = load_test_dataset(predict_filename)
    if ret_code != SUCCESS:
        ret_info['errorCode'] = ret_code
        ret_info['errorMsg'] = CODE_INFO[ret_code]
        return ret_info

    # evaluation
    # correct_sum, predict_sum, recall_sum, recall_correct_sum = 0.0, 0.0, 0.0, 0.0
    # predict: {correct_sum:0.0, predict_sum:0.0, recall_sum:0.0, recall_correct_sum:0.0}
    all_sum = {}
    for sent in golden_dict:
        # 去重
        golden_spo_list = del_duplicate(golden_dict[sent], alias_dict)
        #
        predict_spo_list = predict_result.get(sent, list())
        normalized_predict_spo = del_duplicate(predict_spo_list, alias_dict)
        golden_spo_dic, p_list = list2dic(golden_spo_list)
        predict_dic, n_list = list2dic(normalized_predict_spo)
        for p_l in p_list:
            if p_l not in all_sum:
                all_sum[p_l] = {'correct_sum': 0.0, 'predict_sum': 0.0, 'recall_sum': 0.0,
                                    'recall_correct_sum': 0.0}
        for p_l in n_list:
            if p_l not in all_sum:
                all_sum[p_l] = {'correct_sum': 0.0, 'predict_sum': 0.0, 'recall_sum': 0.0,
                                    'recall_correct_sum': 0.0}
        for key in all_sum.keys():
            golden_value = golden_spo_dic.get(key, list())
            normalized_value = predict_dic.get(key, list())
            all_sum[key]['recall_sum'] += len(golden_value)
            all_sum[key]['predict_sum'] += len(normalized_value)
            for spo in normalized_value:
                if is_spo_in_list(spo, golden_value, alias_dict):
                    all_sum[key]['correct_sum'] += 1
            for golden_spo in golden_value:
                if is_spo_in_list(golden_spo, normalized_value, alias_dict):
                    all_sum[key]['recall_correct_sum'] += 1
    all_info = []
    for key in list(all_sum.keys()):
        correct_sum = all_sum[key]['correct_sum']
        predict_sum = all_sum[key]['predict_sum']
        recall_sum = all_sum[key]['recall_sum']
        recall_correct_sum = all_sum[key]['recall_correct_sum']
        sys.stderr.write(f'{key} correct spo num = {correct_sum}\n')
        sys.stderr.write(f'{key} submitted spo num = {predict_sum}\n')
        sys.stderr.write(f'{key} golden set spo num = {recall_sum}\n')
        sys.stderr.write(f'{key} submitted recall spo num = {recall_correct_sum}\n')
        precision = correct_sum / predict_sum if predict_sum > 0 else 0.0
        recall = recall_correct_sum / recall_sum if recall_sum > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) \
            if precision + recall > 0 else 0.0
        precision = round(precision, 4)
        recall = round(recall, 4)
        f1 = round(f1, 4)
        ret_info = {}
        # ret_info['errorCode'] = SUCCESS
        # ret_info['errorMsg'] = CODE_INFO[SUCCESS]
        ret_info['key'] = key
        ret_info['correct_sum'] = correct_sum
        ret_info['predict_sum'] = predict_sum
        ret_info['recall_sum'] = recall_sum
        ret_info['recall_correct_sum'] = recall_correct_sum
        ret_info['precision'] = precision
        ret_info['recall'] = recall
        ret_info['f1-score'] = f1
        all_info.append(ret_info)
    return all_info

def read_csv():
    file_path = r'E:\文档\4.郑州大学_研究生\60.chip2020评测二\150.csv'
    t_dic = {}
    with open(file_path, encoding="utf-8") as fr:
        next(fr)
        for line in fr:
            line = line.strip().replace('\"','').split(",")
            file_name = line[0].split("/")[-1]
            t_dic[file_name] = line[1:]
    return t_dic

def chip2020_test2():
    #
    t_dic = read_csv()
    #
    s_dic = {}
    # test2数据所在的文件夹
    file_path = r'E:\文档\4.郑州大学_研究生\60.chip2020评测二\150\150'
    file_lst = os.listdir(file_path)
    # 保存至workbook中
    wb = Workbook()
    ws = wb.create_sheet("整体排行榜")
    ws1 = wb.create_sheet("队伍成绩排行榜")
    ws.append(["content","submission_id","team_id","team_name","submit_datetime","file_name","user","real_name","organization","correct_sum", "predict_sum", "recall_sum", "recall_correct_sum", "precision", "recall", "F1"])
    for file_name in file_lst:
        predict_filename = os.path.join(file_path, file_name)
        golden_filename = r"E:\pythonWorkSpace\3.语料处理\面向医学文本的医疗实体和关系抽取评测任务\cmeie_standard_dataset\cmeie_answer\test2_answer.json"
        # alias_filename = r'E:\pythonWorkSpace\3.语料处理\面向医学文本的医疗实体和关系抽取评测任务\cmeie_standard_dataset\cmeie_answer\alias.txt'
        alias_filename = ''
        # print(file_name)
        try:
            ret_2 = calc_pr(predict_filename, alias_filename, golden_filename)
            ans = [file_name]+ t_dic[file_name] + ret_2[1:]
        except:
            ret_2 = [0,0,0,0,0,0,0,0]
            ans = [file_name] + t_dic[file_name] + ret_2[1:]
        if ans[2] not in s_dic:
            s_dic[ans[2]] = {}
            s_dic[ans[2]]["f1"] = 0
        if ans[2] in s_dic and ans[-1] > s_dic[ans[2]]["f1"]:
            s_dic[ans[2]]["f1"] = ans[-1]
            s_dic[ans[2]]["ans"] = ans
        ws.append(ans)
    ws1.append(["content","submission_id","team_id","team_name","submit_datetime","file_name","user","real_name","organization","correct_sum", "predict_sum", "recall_sum", "recall_correct_sum", "precision", "recall", "F1"])
    for k in s_dic.keys():
        ws1.append(s_dic[k]["ans"])
    wb.save("中文医学文本实体关系抽取_test2排行榜.xlsx")

def chip2020_single():
    # 数据所在的文件夹
    file_path = r'C:\Users\guantf\Desktop\CHIP 2020 scsn_dango test2 结果-v2'
    file_lst = os.listdir(file_path)
    #
    wb = Workbook()
    ws = wb.create_sheet("test2得分")
    ws.append(["file_name", "correct_sum", "predict_sum", "recall_sum", "recall_correct_sum", "precision", "recall", "F1"])
    for file_name in file_lst:
        predict_filename = os.path.join(file_path, file_name)
        golden_filename = r"E:\pythonWorkSpace\3.语料处理\面向医学文本的医疗实体和关系抽取评测任务\cmeie_standard_dataset\cmeie_answer\test2_answer.json"
        # alias_filename = r'E:\pythonWorkSpace\3.语料处理\面向医学文本的医疗实体和关系抽取评测任务\cmeie_standard_dataset\cmeie_answer\alias.txt'
        alias_filename = ''
        ret_2 = calc_pr(predict_filename, alias_filename, golden_filename)
        ans = [file_name] + ret_2[1:]
        ws.append(ans)

    wb.save("test2得分.xlsx")
        # print(ans)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--golden_file", type=str, help="true spo results", required=True)
    parser.add_argument(
        "--predict_file", type=str, help="spo results predicted", required=True)
    parser.add_argument(
        "--alias_file", type=str, default='', help="entities alias dictionary")
    args = parser.parse_args()
    golden_filename = args.golden_file
    predict_filename = args.predict_file
    alias_filename = args.alias_file
    ret_2 = calc_pr(predict_filename, alias_filename, golden_filename)
    print(ret_2)
    print('finish!!!')
